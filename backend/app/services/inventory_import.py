"""Inventory valuation import service — Excel「4.3期末全在庫」シートから在庫数量を取り込み、
標準単価×数量で在庫評価金額を算出する。

Excelシート構造:
  A列: 商品コード, C列: 倉庫名, D列: 商品名, F列: 単位名, G列: 当月在庫数,
  H列: 商品区分名, L列: 単価, M列: 金額, O列: 在庫計上

標準単価の取得:
  - 製品/半製品/商品 → StandardCost.unit_cost (39期1月)
  - 原体              → CrudeProductStandardCost.unit_cost (39期1月)
  - 原材料/副資材     → Material.standard_unit_price
  - L列に値がある場合はそれを優先（資材在庫の標準単価が直接設定されているケース）
"""

import io
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit import ImportBatch, ImportError as ImportErrorModel, ImportStatus
from app.models.cost import (
    CrudeProductStandardCost,
    InventoryCategory,
    InventoryValuation,
    SourceSystem,
    StandardCost,
)
from app.models.master import CrudeProduct, FiscalPeriod, Material, Product


# 「商品区分名」(H列) → InventoryCategory のマッピング
CATEGORY_MAP: dict[str, InventoryCategory] = {
    "製品": InventoryCategory.product,
    "半製品": InventoryCategory.semi_finished,
    "原体": InventoryCategory.crude_product,
    "原液": InventoryCategory.crude_product,
    "原材料": InventoryCategory.raw_material,
    "副資材": InventoryCategory.sub_material,
    "商品": InventoryCategory.merchandise,
}

INVENTORY_SHEET_NAME = "4.3期末全在庫"


def _parse_inventory_xlsx(content: bytes, sheet_name: str = INVENTORY_SHEET_NAME) -> list[dict]:
    """4.3期末全在庫シートをパース。1行目=ヘッダ、2行目以降=データ。"""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # シート名一致で見つからない場合、active を使う
        ws = wb.active

    rows: list[dict] = []
    for r_idx, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            continue  # header
        # A=0, C=2, D=3, F=5, G=6, H=7, L=11, M=12, O=14
        if not row_values or row_values[0] is None:
            continue
        item_code = str(row_values[0]).strip() if row_values[0] is not None else ""
        if not item_code:
            continue
        rows.append({
            "row_number": r_idx,
            "item_code": item_code,
            "warehouse_name": (str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] is not None else ""),
            "item_name": (str(row_values[3]).strip() if len(row_values) > 3 and row_values[3] is not None else None),
            "unit": (str(row_values[5]).strip() if len(row_values) > 5 and row_values[5] is not None else "個"),
            "quantity_raw": row_values[6] if len(row_values) > 6 else None,
            "category_raw": (str(row_values[7]).strip() if len(row_values) > 7 and row_values[7] is not None else ""),
            "excel_unit_price": row_values[11] if len(row_values) > 11 else None,
            "excel_amount": row_values[12] if len(row_values) > 12 else None,
        })
    wb.close()
    return rows


def _to_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


async def _build_lookups(
    db: AsyncSession, period_id: uuid.UUID
) -> dict[str, dict]:
    """商品コードからマスタ・標準単価を引くための辞書を構築する。"""
    # Product: code → (id, ProductType)
    result = await db.execute(select(Product.code, Product.id))
    product_map = {row[0]: row[1] for row in result.all()}

    # CrudeProduct: code → id
    result = await db.execute(select(CrudeProduct.code, CrudeProduct.id))
    crude_map = {row[0]: row[1] for row in result.all()}

    # Material: code → (id, standard_unit_price)
    result = await db.execute(
        select(Material.code, Material.id, Material.standard_unit_price)
    )
    material_map = {row[0]: (row[1], row[2]) for row in result.all()}

    # StandardCost (39期1月など、period_id の値で引く): product_id → unit_cost
    result = await db.execute(
        select(StandardCost.product_id, StandardCost.unit_cost).where(
            StandardCost.period_id == period_id
        )
    )
    std_cost_map = {row[0]: row[1] for row in result.all()}

    # CrudeProductStandardCost: crude_product_id → unit_cost
    result = await db.execute(
        select(
            CrudeProductStandardCost.crude_product_id,
            CrudeProductStandardCost.unit_cost,
        ).where(CrudeProductStandardCost.period_id == period_id)
    )
    crude_std_map = {row[0]: row[1] for row in result.all()}

    return {
        "product": product_map,
        "crude": crude_map,
        "material": material_map,
        "std_cost": std_cost_map,
        "crude_std": crude_std_map,
    }


def _resolve_category(category_raw: str) -> InventoryCategory:
    return CATEGORY_MAP.get(category_raw, InventoryCategory.other)


def _resolve_master_and_unit_price(
    item_code: str,
    category: InventoryCategory,
    excel_unit_price: Decimal | None,
    lookups: dict,
) -> tuple[uuid.UUID | None, uuid.UUID | None, uuid.UUID | None, Decimal]:
    """
    item_code・区分から (product_id, crude_product_id, material_id, standard_unit_price) を解決する。
    Excelに単価が直接記載されていれば優先採用、なければマスタから引く。
    マスタ未登録時は (None, None, None, excel単価 or 0) を返す。
    """
    product_id: uuid.UUID | None = None
    crude_id: uuid.UUID | None = None
    material_id: uuid.UUID | None = None
    unit_price: Decimal = Decimal("0")

    if category in (
        InventoryCategory.product,
        InventoryCategory.semi_finished,
        InventoryCategory.merchandise,
    ):
        pid = lookups["product"].get(item_code)
        if pid:
            product_id = pid
            std = lookups["std_cost"].get(pid)
            if std is not None:
                unit_price = Decimal(str(std))
    elif category == InventoryCategory.crude_product:
        cid = lookups["crude"].get(item_code)
        if cid:
            crude_id = cid
            std = lookups["crude_std"].get(cid)
            if std is not None:
                unit_price = Decimal(str(std))
    elif category in (InventoryCategory.raw_material, InventoryCategory.sub_material):
        mat = lookups["material"].get(item_code)
        if mat:
            material_id, mat_price = mat
            unit_price = Decimal(str(mat_price)) if mat_price is not None else Decimal("0")

    # Excel に単価が直接設定されていればそちらを優先（資材在庫の単価設定済みケース）
    if excel_unit_price is not None and excel_unit_price > 0:
        unit_price = excel_unit_price

    return product_id, crude_id, material_id, unit_price


async def process_inventory_import(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    period_id: uuid.UUID,
    sheet_name: str = INVENTORY_SHEET_NAME,
    source_system: str = "manual",
) -> ImportBatch:
    """4.3期末全在庫シートを取り込み、inventory_valuations テーブルにUPSERTする。"""
    batch = ImportBatch(
        file_name=filename,
        source_system=source_system,
        status=ImportStatus.processing,
        period_id=period_id,
        total_rows=0,
        success_rows=0,
        error_rows=0,
        started_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()

    # 期間存在チェック
    period_result = await db.execute(
        select(FiscalPeriod).where(FiscalPeriod.id == period_id)
    )
    period = period_result.scalar_one_or_none()
    if not period:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"指定された会計期間が見つかりません: {period_id}"
        db.add(ImportErrorModel(
            batch_id=batch.id, row_number=0,
            error_message=batch.notes,
        ))
        await db.flush()
        await db.refresh(batch)
        return batch

    # Parse
    try:
        rows = _parse_inventory_xlsx(file_content, sheet_name)
    except Exception as e:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"ファイルパースエラー: {e}"
        db.add(ImportErrorModel(
            batch_id=batch.id, row_number=0,
            error_message=batch.notes,
        ))
        await db.flush()
        await db.refresh(batch)
        return batch

    batch.total_rows = len(rows)
    lookups = await _build_lookups(db, period_id)

    success = 0
    errors = 0
    src_enum = SourceSystem(source_system) if source_system in SourceSystem._value2member_map_ else SourceSystem.manual

    for row in rows:
        try:
            quantity = _to_decimal(row["quantity_raw"]) or Decimal("0")
            excel_price = _to_decimal(row["excel_unit_price"])
            category = _resolve_category(row["category_raw"])

            product_id, crude_id, material_id, unit_price = _resolve_master_and_unit_price(
                row["item_code"], category, excel_price, lookups
            )
            valuation = quantity * unit_price

            # UPSERT (item_code × warehouse × period)
            existing_result = await db.execute(
                select(InventoryValuation).where(
                    InventoryValuation.item_code == row["item_code"],
                    InventoryValuation.warehouse_name == row["warehouse_name"],
                    InventoryValuation.period_id == period_id,
                )
            )
            existing = existing_result.scalar_one_or_none()
            if existing:
                existing.item_name = row["item_name"]
                existing.category = category
                existing.product_id = product_id
                existing.crude_product_id = crude_id
                existing.material_id = material_id
                existing.quantity = quantity
                existing.unit = row["unit"] or "個"
                existing.standard_unit_price = unit_price
                existing.valuation_amount = valuation
                existing.source_system = src_enum
            else:
                db.add(InventoryValuation(
                    period_id=period_id,
                    item_code=row["item_code"],
                    item_name=row["item_name"],
                    warehouse_name=row["warehouse_name"] or "未指定",
                    category=category,
                    product_id=product_id,
                    crude_product_id=crude_id,
                    material_id=material_id,
                    quantity=quantity,
                    unit=row["unit"] or "個",
                    standard_unit_price=unit_price,
                    valuation_amount=valuation,
                    source_system=src_enum,
                ))
            success += 1
        except Exception as e:
            db.add(ImportErrorModel(
                batch_id=batch.id,
                row_number=row.get("row_number", 0),
                error_message=str(e),
                raw_data={k: (str(v) if v is not None else None) for k, v in row.items() if k != "row_number"},
            ))
            errors += 1

    batch.success_rows = success
    batch.error_rows = errors
    batch.status = ImportStatus.completed if errors == 0 else (
        ImportStatus.failed if success == 0 else ImportStatus.completed
    )
    batch.completed_at = datetime.now()

    await db.flush()
    await db.refresh(batch)
    return batch
