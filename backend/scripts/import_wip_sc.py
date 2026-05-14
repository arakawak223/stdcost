"""決算用SC仕掛品.xlsx から仕掛品(半製品) SC単価を取り込む。

取込対象:
  1. シート「仕掛品標準単価一覧表（貼付）」(R3〜R41)
     → wip_standard_costs テーブルに bulk-upsert (period_id 指定)
     列構成: B=仕掛原液(consolidation_key), C=前工程費, D=原材料費,
             E=労務費, F=経費, G=計(unit_cost)
  2. シート「仕掛品名寄（貼付）」(R4〜R472)
     → 番手付きコード→consolidation_key の解決マップとして使用。
     → products テーブルの sc_consolidation_key を更新:
        a. 直接マッチ: products.name が wip_standard_costs のキーと完全一致
        b. 名寄マッチ: products.name が名寄マップのキーと一致
        c. 半製品 (product_type='semi_finished') かつ未マッチ → 'その他'

使い方:
  cd backend
  python -m scripts.import_wip_sc \\
      --period-id 5ce52674-b54a-4ec1-817c-140cbd0dfd80 \\
      --xlsx /workspaces/stdcost/docs/reference/決算用SC仕掛品.xlsx
  # --commit を付けると DB に書き込み (省略時は dry-run)
"""
import argparse
import asyncio
import os
import sys
import uuid
from datetime import date
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from sqlalchemy import select, func

from app.db.session import async_session_factory
from app.models.cost import (
    InventoryCategory,
    InventoryValuation,
    WipStandardCost,
)
from app.models.master import Product, ProductType
from app.services.inventory_valuation import recalculate_valuation_amounts


SHEET_PRICES = "仕掛品標準単価一覧表（貼付）"
SHEET_NAYOSE = "仕掛品名寄（貼付）"


def _to_decimal(v) -> Decimal:
    """Excel セル値を Decimal に変換。'ー' / None / 空文字は 0 とする。"""
    if v is None:
        return Decimal("0")
    s = str(v).strip()
    if not s or s in ("ー", "−", "-"):
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def load_price_table(xlsx_path: str) -> list[dict]:
    """仕掛品標準単価一覧表 シートからキー別単価を取得。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET_PRICES]
    rows = []
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri < 3:
            continue
        # B(1)=key, C(2)=前工程費, D(3)=原材料費, E(4)=労務費, F(5)=経費, G(6)=計
        if len(row) < 7:
            continue
        key = row[1]
        if not key:
            continue
        key_str = str(key).strip()
        if not key_str:
            continue
        rows.append({
            "consolidation_key": key_str,
            "pre_process_cost": _to_decimal(row[2]),
            "material_cost": _to_decimal(row[3]),
            "labor_cost": _to_decimal(row[4]),
            "expense_cost": _to_decimal(row[5]),
            "unit_cost": _to_decimal(row[6]),
        })
    wb.close()
    return rows


def load_nayose_map(xlsx_path: str) -> dict[str, str]:
    """仕掛品名寄 シートから {番手付きコード → consolidation_key} マップを構築。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAYOSE]
    out: dict[str, str] = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri < 4:
            continue
        if len(row) < 3:
            continue
        code = row[1]
        nayose = row[2]
        if not code or not nayose:
            continue
        out[str(code).strip()] = str(nayose).strip()
    wb.close()
    return out


async def upsert_wip_costs(
    db, period_id: uuid.UUID, rows: list[dict], commit: bool
) -> tuple[int, int, int]:
    """wip_standard_costs を bulk upsert。返り値: (inserted, updated, unchanged)"""
    existing_res = await db.execute(
        select(WipStandardCost).where(WipStandardCost.period_id == period_id)
    )
    existing_map: dict[str, WipStandardCost] = {
        r.consolidation_key: r for r in existing_res.scalars().all()
    }

    inserted = 0
    updated = 0
    unchanged = 0
    for item in rows:
        existing = existing_map.get(item["consolidation_key"])
        if existing is None:
            db.add(WipStandardCost(
                consolidation_key=item["consolidation_key"],
                period_id=period_id,
                unit_cost=item["unit_cost"],
                pre_process_cost=item["pre_process_cost"],
                material_cost=item["material_cost"],
                labor_cost=item["labor_cost"],
                expense_cost=item["expense_cost"],
                effective_date=None,
                notes="決算用SC仕掛品.xlsx 仕掛品標準単価一覧表（貼付）",
            ))
            inserted += 1
        else:
            changed = False
            for f in ("unit_cost", "pre_process_cost", "material_cost", "labor_cost", "expense_cost"):
                if Decimal(str(getattr(existing, f))) != item[f]:
                    setattr(existing, f, item[f])
                    changed = True
            if changed:
                updated += 1
            else:
                unchanged += 1

    if commit:
        await db.flush()
    return inserted, updated, unchanged


async def update_product_keys(
    db, valid_keys: set[str], nayose_map: dict[str, str], commit: bool
) -> dict[str, int]:
    """products.sc_consolidation_key を解決して更新。"""
    # 全 products
    res = await db.execute(select(Product))
    products = list(res.scalars().all())

    direct = 0
    via_nayose = 0
    other = 0
    skipped = 0
    changed = 0

    for p in products:
        key: str | None = None
        # 候補値: name / code を順に試す
        for cand in (p.name, p.code):
            if not cand:
                continue
            cand_s = str(cand).strip()
            if cand_s in valid_keys:
                key = cand_s
                direct += 1
                break
            if cand_s in nayose_map:
                resolved = nayose_map[cand_s]
                if resolved in valid_keys:
                    key = resolved
                    via_nayose += 1
                    break
                # 名寄結果が一覧にも未登録 → "その他"
                if resolved == "その他":
                    key = "その他"
                    other += 1
                    break
        if key is None:
            # 半製品で未マッチなら "その他" を割り当て、それ以外はスキップ
            if p.product_type == ProductType.semi_finished:
                key = "その他"
                other += 1
            else:
                skipped += 1
                continue

        if p.sc_consolidation_key != key:
            p.sc_consolidation_key = key
            changed += 1

    if commit:
        await db.flush()
    return {
        "direct": direct,
        "via_nayose": via_nayose,
        "other": other,
        "skipped_non_semi_finished": skipped,
        "changed": changed,
        "total_products": len(products),
    }


async def run(period_id: uuid.UUID, xlsx_path: str, commit: bool):
    print(f"=== Excel 読込 ===")
    price_rows = load_price_table(xlsx_path)
    nayose_map = load_nayose_map(xlsx_path)
    print(f"  price rows: {len(price_rows)}件")
    print(f"  nayose map: {len(nayose_map)}件")
    print(f"  価格キーサンプル: {[r['consolidation_key'] for r in price_rows[:8]]}")

    valid_keys = {r["consolidation_key"] for r in price_rows}

    async with async_session_factory() as db:
        print(f"\n=== wip_standard_costs upsert (period_id={period_id}) ===")
        ins, upd, unc = await upsert_wip_costs(db, period_id, price_rows, commit)
        print(f"  INSERT: {ins}件 / UPDATE: {upd}件 / UNCHANGED: {unc}件")

        print(f"\n=== products.sc_consolidation_key 更新 ===")
        stats = await update_product_keys(db, valid_keys, nayose_map, commit)
        for k, v in stats.items():
            print(f"  {k}: {v}")

        if commit:
            print(f"\n=== inventory_valuation 再計算 (period_id={period_id}) ===")
            updated = await recalculate_valuation_amounts(db, period_id)
            print(f"  updated: {updated}")

            await db.commit()
            print(f"\n=== サマリ (period_id={period_id}) ===")
            for cat in (
                InventoryCategory.semi_finished,
                InventoryCategory.product,
                InventoryCategory.merchandise,
                InventoryCategory.crude_product,
                InventoryCategory.raw_material,
                InventoryCategory.sub_material,
            ):
                cnt_res = await db.execute(
                    select(func.count(), func.sum(InventoryValuation.valuation_amount)).where(
                        InventoryValuation.period_id == period_id,
                        InventoryValuation.category == cat,
                    )
                )
                cnt, total = cnt_res.one()
                priced_res = await db.execute(
                    select(func.count()).where(
                        InventoryValuation.period_id == period_id,
                        InventoryValuation.category == cat,
                        InventoryValuation.standard_unit_price > 0,
                    )
                )
                priced = priced_res.scalar()
                print(f"  {cat.value}: 件数={cnt} priced={priced} total=¥{float(total or 0):,.0f}")
        else:
            print("\n[DRY-RUN] --commit を付与すると DB に書き込みます。")


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--period-id", required=True)
    parser.add_argument(
        "--xlsx",
        default="/workspaces/stdcost/docs/reference/決算用SC仕掛品.xlsx",
    )
    parser.add_argument("--commit", action="store_true")
    args = parser.parse_args()
    os.environ.setdefault("SECRET_KEY", "dev")
    os.environ.setdefault("ANTHROPIC_API_KEY", "dummy")
    asyncio.run(run(uuid.UUID(args.period_id), args.xlsx, args.commit))


if __name__ == "__main__":
    main()
