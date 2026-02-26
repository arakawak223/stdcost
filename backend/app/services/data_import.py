"""Data import service — CSV/Excel ファイルから実際原価データを取り込む。

設計方針: 実際原価は簡素化した集計データとして取り込む。
各ソースシステムのカラムマッピングを定義し、コード→UUID解決を行う。
"""

import csv
import io
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit import ImportBatch, ImportError as ImportErrorModel, ImportStatus
from app.models.cost import ActualCost, CrudeProductActualCost, SourceSystem
from app.models.master import CostCenter, CrudeProduct, Material, Product


# --- Source system column mappings ---

SOURCE_MAPPINGS: dict[str, dict] = {
    "sc_system": {
        "file_type": "csv",
        "encoding": "shift_jis",
        "target_table": "actual_cost",
        "column_map": {
            "品目コード": "product_code",
            "部門コード": "cost_center_code",
            "原体原価": "crude_product_cost",
            "資材費": "packaging_cost",
            "労務費": "labor_cost",
            "経費": "overhead_cost",
            "外注加工費": "outsourcing_cost",
            "合計": "total_cost",
            "生産数量": "quantity_produced",
            "備考": "notes",
        },
    },
    "geneki_db": {
        "file_type": "xlsx",
        "encoding": None,
        "target_table": "crude_product_actual_cost",
        "sheet_name": None,  # first sheet
        "column_map": {
            "原体コード": "crude_product_code",
            "原材料費": "material_cost",
            "労務費": "labor_cost",
            "経費": "overhead_cost",
            "前工程費": "prior_process_cost",
            "合計": "total_cost",
            "実際数量": "actual_quantity",
            "備考": "notes",
        },
    },
    "kanjyo_bugyo": {
        "file_type": "csv",
        "encoding": "shift_jis",
        "target_table": "actual_cost",
        "column_map": {
            "品目コード": "product_code",
            "部門コード": "cost_center_code",
            "原体原価": "crude_product_cost",
            "資材費": "packaging_cost",
            "労務費": "labor_cost",
            "経費": "overhead_cost",
            "外注加工費": "outsourcing_cost",
            "合計": "total_cost",
            "生産数量": "quantity_produced",
            "備考": "notes",
        },
    },
    "manual": {
        "file_type": "csv",
        "encoding": "utf-8-sig",
        "target_table": "actual_cost",
        "column_map": {
            "product_code": "product_code",
            "cost_center_code": "cost_center_code",
            "crude_product_cost": "crude_product_cost",
            "packaging_cost": "packaging_cost",
            "labor_cost": "labor_cost",
            "overhead_cost": "overhead_cost",
            "outsourcing_cost": "outsourcing_cost",
            "total_cost": "total_cost",
            "quantity_produced": "quantity_produced",
            "notes": "notes",
        },
    },
}


async def process_import(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    source_system: str,
    period_id: uuid.UUID,
) -> ImportBatch:
    """ファイルをパースし、実際原価データを取り込む。"""

    # Create batch record
    batch = ImportBatch(
        file_name=filename,
        source_system=source_system,
        status=ImportStatus.processing,
        period_id=period_id,
        total_rows=0,
        success_rows=0,
        error_rows=0,
        started_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()

    mapping = SOURCE_MAPPINGS[source_system]
    target_table = mapping["target_table"]
    column_map = mapping["column_map"]

    # Parse file
    try:
        if mapping["file_type"] == "csv":
            rows = _parse_csv(file_content, mapping["encoding"], column_map)
        else:
            sheet_name = mapping.get("sheet_name")
            rows = _parse_xlsx(file_content, sheet_name, column_map)
    except Exception as e:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"ファイルパースエラー: {e}"
        error = ImportErrorModel(
            batch_id=batch.id,
            row_number=0,
            error_message=f"ファイルパースエラー: {e}",
        )
        db.add(error)
        await db.flush()
        await db.refresh(batch)
        return batch

    batch.total_rows = len(rows)

    # Build code→UUID lookups
    lookups = await _build_code_lookup(db)

    # Process rows
    success_count = 0
    error_count = 0

    for i, row in enumerate(rows, start=2):  # row 1 = header
        error_msg = _validate_and_transform(row, lookups, target_table)
        if error_msg:
            error = ImportErrorModel(
                batch_id=batch.id,
                row_number=i,
                error_message=error_msg,
                raw_data=row,
            )
            db.add(error)
            error_count += 1
            continue

        try:
            await _upsert_record(db, target_table, row, period_id, source_system)
            success_count += 1
        except Exception as e:
            error = ImportErrorModel(
                batch_id=batch.id,
                row_number=i,
                error_message=str(e),
                raw_data=row,
            )
            db.add(error)
            error_count += 1

    batch.success_rows = success_count
    batch.error_rows = error_count
    batch.status = ImportStatus.completed if error_count == 0 else ImportStatus.completed
    batch.completed_at = datetime.now()

    if error_count > 0 and success_count == 0:
        batch.status = ImportStatus.failed

    await db.flush()
    await db.refresh(batch)
    return batch


def _parse_csv(
    content: bytes, encoding: str, column_map: dict[str, str]
) -> list[dict[str, str]]:
    """CSV ファイルをパースし、カラムマッピング適用済みの行リストを返す。"""
    text = content.decode(encoding)
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, str]] = []
    reverse_map = {src: dst for src, dst in column_map.items()}

    for raw_row in reader:
        mapped: dict[str, str] = {}
        for src_col, dst_col in reverse_map.items():
            value = raw_row.get(src_col, "").strip()
            if value:
                mapped[dst_col] = value
        rows.append(mapped)

    return rows


def _parse_xlsx(
    content: bytes, sheet_name: str | None, column_map: dict[str, str]
) -> list[dict[str, str]]:
    """Excel (.xlsx) ファイルをパースし、カラムマッピング適用済みの行リストを返す。"""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    headers = [str(cell) if cell else "" for cell in next(rows_iter)]
    reverse_map = {src: dst for src, dst in column_map.items()}

    rows: list[dict[str, str]] = []
    for row_values in rows_iter:
        mapped: dict[str, str] = {}
        for idx, cell_value in enumerate(row_values):
            if idx < len(headers):
                header = headers[idx]
                if header in reverse_map and cell_value is not None:
                    mapped[reverse_map[header]] = str(cell_value).strip()
        if mapped:  # skip entirely empty rows
            rows.append(mapped)

    wb.close()
    return rows


async def _build_code_lookup(db: AsyncSession) -> dict[str, dict[str, uuid.UUID]]:
    """マスタテーブルから code→UUID の辞書を構築する。"""
    lookups: dict[str, dict[str, uuid.UUID]] = {}

    # Products
    result = await db.execute(select(Product.code, Product.id))
    lookups["product"] = {row[0]: row[1] for row in result.all()}

    # CrudeProducts
    result = await db.execute(select(CrudeProduct.code, CrudeProduct.id))
    lookups["crude_product"] = {row[0]: row[1] for row in result.all()}

    # CostCenters
    result = await db.execute(select(CostCenter.code, CostCenter.id))
    lookups["cost_center"] = {row[0]: row[1] for row in result.all()}

    # Materials
    result = await db.execute(select(Material.code, Material.id))
    lookups["material"] = {row[0]: row[1] for row in result.all()}

    return lookups


def _validate_and_transform(
    row: dict[str, str],
    lookups: dict[str, dict[str, uuid.UUID]],
    target_table: str,
) -> str | None:
    """行データをバリデーションし、コード→UUIDに変換する。エラー時はメッセージを返す。"""

    if target_table == "actual_cost":
        # product_code → product_id
        product_code = row.get("product_code", "")
        if not product_code:
            return "品目コードが空です"
        product_id = lookups["product"].get(product_code)
        if not product_id:
            return f"品目コード '{product_code}' がマスタに存在しません"
        row["product_id"] = product_id
        del row["product_code"]

        # cost_center_code → cost_center_id
        cc_code = row.get("cost_center_code", "")
        if not cc_code:
            return "部門コードが空です"
        cc_id = lookups["cost_center"].get(cc_code)
        if not cc_id:
            return f"部門コード '{cc_code}' がマスタに存在しません"
        row["cost_center_id"] = cc_id
        del row["cost_center_code"]

    elif target_table == "crude_product_actual_cost":
        # crude_product_code → crude_product_id
        cp_code = row.get("crude_product_code", "")
        if not cp_code:
            return "原体コードが空です"
        cp_id = lookups["crude_product"].get(cp_code)
        if not cp_id:
            return f"原体コード '{cp_code}' がマスタに存在しません"
        row["crude_product_id"] = cp_id
        del row["crude_product_code"]

    # Convert numeric fields to Decimal
    numeric_fields = [
        "crude_product_cost", "packaging_cost", "labor_cost", "overhead_cost",
        "outsourcing_cost", "total_cost", "quantity_produced",
        "material_cost", "prior_process_cost", "actual_quantity",
    ]
    for field in numeric_fields:
        if field in row:
            try:
                row[field] = Decimal(str(row[field]))
            except (InvalidOperation, ValueError):
                return f"'{field}' の値 '{row[field]}' が数値として不正です"

    return None


async def _upsert_record(
    db: AsyncSession,
    target_table: str,
    data: dict,
    period_id: uuid.UUID,
    source_system: str,
) -> None:
    """既存レコードがあれば更新、なければ作成する。"""
    # Remove non-model fields
    clean = {k: v for k, v in data.items() if v is not None and v != ""}

    if target_table == "actual_cost":
        product_id = clean["product_id"]
        cost_center_id = clean["cost_center_id"]

        result = await db.execute(
            select(ActualCost).where(
                ActualCost.product_id == product_id,
                ActualCost.cost_center_id == cost_center_id,
                ActualCost.period_id == period_id,
            )
        )
        existing = result.scalar_one_or_none()

        if existing:
            for field, value in clean.items():
                if field not in ("product_id", "cost_center_id"):
                    setattr(existing, field, value)
            existing.source_system = SourceSystem(source_system)
        else:
            record = ActualCost(
                period_id=period_id,
                source_system=SourceSystem(source_system),
                **clean,
            )
            db.add(record)

    elif target_table == "crude_product_actual_cost":
        crude_product_id = clean["crude_product_id"]

        result = await db.execute(
            select(CrudeProductActualCost).where(
                CrudeProductActualCost.crude_product_id == crude_product_id,
                CrudeProductActualCost.period_id == period_id,
            )
        )
        existing = result.scalar_one_or_none()

        if existing:
            for field, value in clean.items():
                if field != "crude_product_id":
                    setattr(existing, field, value)
            existing.source_system = SourceSystem(source_system)
        else:
            record = CrudeProductActualCost(
                period_id=period_id,
                source_system=SourceSystem(source_system),
                **clean,
            )
            db.add(record)

    await db.flush()
