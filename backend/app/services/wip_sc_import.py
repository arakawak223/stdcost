"""仕掛品(半製品) SC 単価インポートサービス —
xlsx「決算用SC仕掛品.xlsx」から仕掛品の標準単価を取り込み、
wip_standard_costs テーブルに upsert + products.sc_consolidation_key を解決する。

シート構造:
  [仕掛品標準単価一覧表（貼付）] (R3〜R41 = 39件)
    B列: 仕掛原液 (consolidation_key, 例 "B" "BM" "FB" "その他")
    C列: 前工程費 (¥/kg)
    D列: 原材料費 (¥/kg)
    E列: 労務費 (¥/kg)
    F列: 経費 (¥/kg)
    G列: 計 (unit_cost ¥/kg)

  [仕掛品名寄（貼付）] (R4〜R472 = 469件)
    B列: 番手付きコード (products.name もしくは products.code とマッチング対象)
    C列: 名寄結果(consolidation_key、未確定なら "その他")

ロジック:
  1. 単価表を読み wip_standard_costs にUPSERT (UNIQUE: consolidation_key × period_id)
  2. 名寄表を読み 番手コード→キー マップを構築
  3. products を走査し、各製品の sc_consolidation_key を解決:
     a. name/code が単価表キー集合に直接マッチ → そのキー
     b. name/code が名寄表に存在し、結果が単価表に存在 → 解決後のキー
     c. 半製品(product_type=semi_finished)で未マッチ → "その他"
     d. 上記いずれでもない (=製品/原体/原材料/商品など) → 何もしない
  4. recalculate_valuation_amounts を呼んで在庫評価金額を再計算
"""

import io
import uuid
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit import ImportBatch, ImportError as ImportErrorModel, ImportStatus
from app.models.cost import SourceSystem, WipStandardCost
from app.models.master import FiscalPeriod, Product, ProductType
from app.services.inventory_valuation import recalculate_valuation_amounts


SHEET_PRICES = "仕掛品標準単価一覧表（貼付）"
SHEET_NAYOSE = "仕掛品名寄（貼付）"


def _to_decimal(v) -> Decimal:
    """Excel セル値を Decimal に変換。'ー' / None / 空文字は 0 とする。"""
    if v is None:
        return Decimal("0")
    s = str(v).strip()
    if not s or s in ("ー", "−", "-"):
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def parse_price_table(content: bytes, sheet_name: str = SHEET_PRICES) -> list[dict]:
    """仕掛品標準単価一覧表 シートからキー別単価を取得。"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows: list[dict] = []
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri < 3:
            continue
        if len(row) < 7:
            continue
        key = row[1]
        if not key:
            continue
        key_str = str(key).strip()
        if not key_str:
            continue
        rows.append({
            "consolidation_key": key_str,
            "pre_process_cost": _to_decimal(row[2]),
            "material_cost": _to_decimal(row[3]),
            "labor_cost": _to_decimal(row[4]),
            "expense_cost": _to_decimal(row[5]),
            "unit_cost": _to_decimal(row[6]),
        })
    wb.close()
    return rows


def parse_nayose_map(content: bytes, sheet_name: str = SHEET_NAYOSE) -> dict[str, str]:
    """仕掛品名寄 シートから {番手付きコード → consolidation_key} マップを構築。"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name]
    out: dict[str, str] = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri < 4:
            continue
        if len(row) < 3:
            continue
        code = row[1]
        nayose = row[2]
        if not code or not nayose:
            continue
        out[str(code).strip()] = str(nayose).strip()
    wb.close()
    return out


async def upsert_wip_costs(
    db: AsyncSession,
    period_id: uuid.UUID,
    rows: list[dict],
    notes_default: str = "決算用SC仕掛品.xlsx",
) -> dict[str, int]:
    """wip_standard_costs を bulk upsert。"""
    existing_res = await db.execute(
        select(WipStandardCost).where(WipStandardCost.period_id == period_id)
    )
    existing_map: dict[str, WipStandardCost] = {
        r.consolidation_key: r for r in existing_res.scalars().all()
    }

    inserted = 0
    updated = 0
    unchanged = 0
    breakdown_fields = ("pre_process_cost", "material_cost", "labor_cost", "expense_cost")
    for item in rows:
        existing = existing_map.get(item["consolidation_key"])
        if existing is None:
            db.add(WipStandardCost(
                consolidation_key=item["consolidation_key"],
                period_id=period_id,
                unit_cost=item["unit_cost"],
                pre_process_cost=item["pre_process_cost"],
                material_cost=item["material_cost"],
                labor_cost=item["labor_cost"],
                expense_cost=item["expense_cost"],
                effective_date=None,
                notes=notes_default,
            ))
            inserted += 1
        else:
            changed = False
            if Decimal(str(existing.unit_cost)) != item["unit_cost"]:
                existing.unit_cost = item["unit_cost"]
                changed = True
            for f in breakdown_fields:
                if Decimal(str(getattr(existing, f))) != item[f]:
                    setattr(existing, f, item[f])
                    changed = True
            if changed:
                updated += 1
            else:
                unchanged += 1
    await db.flush()
    return {"inserted": inserted, "updated": updated, "unchanged": unchanged}


async def update_product_keys(
    db: AsyncSession,
    valid_keys: set[str],
    nayose_map: dict[str, str],
) -> dict[str, int]:
    """products.sc_consolidation_key を解決して更新。"""
    res = await db.execute(select(Product))
    products = list(res.scalars().all())

    direct = 0
    via_nayose = 0
    other = 0
    skipped = 0
    changed = 0

    for p in products:
        key: str | None = None
        for cand in (p.name, p.code):
            if not cand:
                continue
            cand_s = str(cand).strip()
            if cand_s in valid_keys:
                key = cand_s
                direct += 1
                break
            if cand_s in nayose_map:
                resolved = nayose_map[cand_s]
                if resolved in valid_keys:
                    key = resolved
                    via_nayose += 1
                    break
                if resolved == "その他":
                    key = "その他"
                    other += 1
                    break
        if key is None:
            if p.product_type == ProductType.semi_finished:
                key = "その他"
                other += 1
            else:
                skipped += 1
                continue
        if p.sc_consolidation_key != key:
            p.sc_consolidation_key = key
            changed += 1

    await db.flush()
    return {
        "direct": direct,
        "via_nayose": via_nayose,
        "other": other,
        "skipped_non_semi_finished": skipped,
        "changed": changed,
        "total_products": len(products),
    }


async def process_wip_sc_import(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    period_id: uuid.UUID,
    price_sheet: str = SHEET_PRICES,
    nayose_sheet: str = SHEET_NAYOSE,
    source_system: str = "manual",
) -> ImportBatch:
    """仕掛品 SC 単価 Excel を取り込み、ImportBatch を返す。

    フロー:
      1. 単価表シートをパース → wip_standard_costs upsert
      2. 名寄シートをパース → products.sc_consolidation_key 解決
      3. recalculate_valuation_amounts で在庫評価を再計算
    """
    batch = ImportBatch(
        file_name=filename,
        source_system=source_system if source_system in SourceSystem._value2member_map_
        else SourceSystem.manual.value,
        status=ImportStatus.processing,
        period_id=period_id,
        total_rows=0,
        success_rows=0,
        error_rows=0,
        started_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()

    # 期間確認
    period_res = await db.execute(
        select(FiscalPeriod).where(FiscalPeriod.id == period_id)
    )
    if period_res.scalar_one_or_none() is None:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"指定された会計期間が見つかりません: {period_id}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    # Excel パース
    try:
        price_rows = parse_price_table(file_content, price_sheet)
        nayose_map = parse_nayose_map(file_content, nayose_sheet)
    except Exception as e:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"ファイルパースエラー: {e}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    batch.total_rows = len(price_rows)

    valid_keys = {r["consolidation_key"] for r in price_rows}

    # 単価表 upsert
    cost_stats = await upsert_wip_costs(db, period_id, price_rows)
    # products キー解決
    key_stats = await update_product_keys(db, valid_keys, nayose_map)
    # 在庫評価再計算
    revalued = await recalculate_valuation_amounts(db, period_id)

    batch.success_rows = cost_stats["inserted"] + cost_stats["updated"]
    batch.error_rows = 0
    batch.status = ImportStatus.completed
    batch.completed_at = datetime.now()
    batch.notes = (
        f"price rows: {len(price_rows)} (INSERT {cost_stats['inserted']} / "
        f"UPDATE {cost_stats['updated']} / UNCHANGED {cost_stats['unchanged']}); "
        f"products keyed: direct={key_stats['direct']} "
        f"via_nayose={key_stats['via_nayose']} other={key_stats['other']} "
        f"changed={key_stats['changed']} "
        f"(skip non-semi={key_stats['skipped_non_semi_finished']}, "
        f"total={key_stats['total_products']}); "
        f"inventory revalued: {revalued}"
    )
    await db.flush()
    await db.refresh(batch)
    return batch
