"""原材料在庫インポートサービス —
xlsx「決算用SC原材料.xlsx」(または同等構造ファイル) から原材料の
ロット別在庫(1.5原材料在庫シート)を取り込み、
inventory_valuations に category=raw_material として登録する。

シート構造:
  [1.5原材料在庫] (ロット別1143行)
    A列: 原料コード (例 "001", "201", "016-1") — 0埋めあり
    B列: 原料名 (例 "ﾘﾝｺﾞ", "黒糖")
    C列: 原料ロット番号
    D列: 在庫数量
    E列: 単位 ("㎏" or "kg")
    F列: 受入数量 / G列: 使用数量

  [原材料SC明細] (Row 5+ がデータ)
    A列(1): コード (整数表現)
    B列(2): 原料名
    C列(3): SC単価 (円/kg)

ロジック:
  1. 1.5シートをロットコードで集約 → 正規化コード(0埋め除去)に集計
  2. SC明細から (正規化コード → SC単価) のマップを構築
  3. materials マスタを 正規化コード で lookup; 未登録は INSERT (raw type)
  4. SC明細の単価で materials.standard_unit_price を更新 (マスタ最新化)
  5. inventory_valuations に raw_material としてUPSERT
     standard_unit_price = SC単価 (なければ0)
     warehouse_name = "原料倉庫" (固定)
"""

import io
import re
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit import ImportBatch, ImportError as ImportErrorModel, ImportStatus
from app.models.cost import (
    InventoryCategory,
    InventoryValuation,
    MaterialStandardCost,
    SourceSystem,
)
from app.models.master import FiscalPeriod, Material, MaterialType


RAW_INVENTORY_SHEET = "1.5原材料在庫"
RAW_SC_SHEET = "原材料SC明細"
DEFAULT_WAREHOUSE = "原料倉庫"


def _normalize_code(code) -> str:
    """先頭0除去。016-1 のようなハイフン付きはセグメント単位に正規化。
    materials.code 側も同方式で正規化してマッチング。
    """
    s = str(code).strip()
    if not s:
        return ""
    if s.isdigit():
        return s.lstrip("0") or "0"
    if "-" in s:
        parts = s.split("-")
        norm = []
        for p in parts:
            if p.isdigit():
                norm.append(p.lstrip("0") or "0")
            else:
                norm.append(p)
        return "-".join(norm)
    return s


def _to_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _parse_inventory_lots(content: bytes, sheet_name: str = RAW_INVENTORY_SHEET) -> dict[str, dict]:
    """1.5原材料在庫 シートをロット集約してコード単位の在庫を返す。
    返り値: 正規化コード → {orig_code, name, unit, stock_sum, lot_count}
    """
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    agg: dict[str, dict] = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri == 1:
            continue
        if not row or row[0] is None:
            continue
        orig = str(row[0]).strip()
        if not orig:
            continue
        norm = _normalize_code(orig)
        if not norm:
            continue
        d = agg.setdefault(norm, {
            "orig_code": orig, "name": None, "unit": "kg",
            "stock_sum": Decimal("0"), "lot_count": 0,
        })
        if d["name"] is None and row[1] is not None:
            d["name"] = str(row[1]).strip()
        if len(row) > 4 and row[4] is not None:
            u = str(row[4]).strip()
            if u in ("㎏", "kg"):
                d["unit"] = "kg"
            elif u:
                d["unit"] = u[:10]
        stock = _to_decimal(row[3] if len(row) > 3 else None)
        if stock is not None:
            d["stock_sum"] += stock
        d["lot_count"] += 1
    wb.close()
    return agg


def _parse_sc_map(content: bytes, sheet_name: str = RAW_SC_SHEET) -> dict[str, tuple[str | None, Decimal]]:
    """原材料SC明細 シート Row 5+ から 正規化コード → (name, SC単価) を返す。"""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]
    sc_map: dict[str, tuple[str | None, Decimal]] = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri < 5:
            continue
        code_v = row[0] if len(row) > 0 else None
        sc_v = row[2] if len(row) > 2 else None
        if code_v is None or sc_v is None:
            continue
        if isinstance(code_v, (int, float)):
            c_str = str(int(code_v)) if code_v == int(code_v) else str(code_v)
        else:
            c_str = str(code_v).strip()
        norm = _normalize_code(c_str)
        if not norm:
            continue
        try:
            price = Decimal(str(sc_v))
        except (InvalidOperation, ValueError):
            continue
        if price <= 0:
            continue
        if norm not in sc_map:
            sc_map[norm] = (row[1] if len(row) > 1 else None, price)
    wb.close()
    return sc_map


async def process_raw_material_inventory_import(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    period_id: uuid.UUID,
    inventory_sheet: str = RAW_INVENTORY_SHEET,
    sc_sheet: str = RAW_SC_SHEET,
    source_system: str = "manual",
    warehouse_name: str = DEFAULT_WAREHOUSE,
    delete_existing: bool = True,
    skip_zero_stock: bool = False,
    update_master_price: bool = True,
) -> ImportBatch:
    """1.5原材料在庫 シートを取り込み、inventory_valuations にUPSERT。

    update_master_price=True の場合、SC明細の単価で
    materials.standard_unit_price も更新する(マスタ最新化)。
    """
    batch = ImportBatch(
        file_name=filename,
        source_system=source_system,
        status=ImportStatus.processing,
        period_id=period_id,
        total_rows=0,
        success_rows=0,
        error_rows=0,
        started_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()

    period_result = await db.execute(
        select(FiscalPeriod).where(FiscalPeriod.id == period_id)
    )
    period = period_result.scalar_one_or_none()
    if not period:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"指定された会計期間が見つかりません: {period_id}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    # Parse
    try:
        agg = _parse_inventory_lots(file_content, inventory_sheet)
        sc_map = _parse_sc_map(file_content, sc_sheet)
    except Exception as e:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"ファイルパースエラー: {e}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    batch.total_rows = len(agg)

    # 既存削除 (period × category=raw_material × warehouse)
    if delete_existing:
        from sqlalchemy import delete
        await db.execute(
            delete(InventoryValuation).where(
                InventoryValuation.period_id == period_id,
                InventoryValuation.category == InventoryCategory.raw_material,
                InventoryValuation.warehouse_name == warehouse_name,
            )
        )

    # materials マスタ正規化マップ (正規化code → (id, orig_code, name, current_price))
    res = await db.execute(
        select(Material.id, Material.code, Material.name, Material.standard_unit_price)
    )
    mat_map: dict[str, tuple[uuid.UUID, str, str, Decimal]] = {}
    for r in res.all():
        norm = _normalize_code(r[1])
        mat_map[norm] = (r[0], r[1], r[2], r[3])

    src_enum = SourceSystem(source_system) if source_system in SourceSystem._value2member_map_ else SourceSystem.manual

    success = 0
    errors = 0
    skipped_zero = 0
    new_mat_count = 0
    sc_priced = 0
    no_sc_count = 0
    master_updated = 0

    # マスタ単価更新マップ (正規化code → 新価格)
    pending_master_updates: dict[uuid.UUID, Decimal] = {}

    for norm_code, d in agg.items():
        try:
            stock = d["stock_sum"]
            if abs(stock) < Decimal("0.001"):
                stock = Decimal("0")
            if skip_zero_stock and stock == 0:
                skipped_zero += 1
                continue

            # materials lookup or create
            sc_entry = sc_map.get(norm_code)
            sc_price = sc_entry[1] if sc_entry else None

            if norm_code in mat_map:
                mat_id, mat_orig_code, mat_name, current_price = mat_map[norm_code]
            else:
                # materials マスタに INSERT
                new_mat = Material(
                    code=d["orig_code"][:20],  # 元の0埋めコードを保持
                    name=(d.get("name") or norm_code)[:200],
                    material_type=MaterialType.raw,
                    unit=d.get("unit") or "kg",
                    standard_unit_price=sc_price if sc_price else Decimal("0"),
                    is_active=True,
                    notes="1.5原材料在庫 取込時に補完",
                )
                db.add(new_mat)
                await db.flush()
                mat_id = new_mat.id
                mat_orig_code = d["orig_code"]
                current_price = sc_price or Decimal("0")
                mat_map[norm_code] = (mat_id, mat_orig_code, d.get("name"), current_price)
                new_mat_count += 1

            # マスタ単価更新 (差異がある場合のみ)
            if update_master_price and sc_price is not None and Decimal(str(current_price)) != sc_price:
                pending_master_updates[mat_id] = sc_price

            unit_price = sc_price if sc_price is not None else Decimal("0")
            if sc_price is not None:
                sc_priced += 1
            else:
                no_sc_count += 1

            valuation = stock * unit_price

            # UPSERT inventory_valuations (item_code = orig_code)
            existing_result = await db.execute(
                select(InventoryValuation).where(
                    InventoryValuation.item_code == d["orig_code"],
                    InventoryValuation.warehouse_name == warehouse_name,
                    InventoryValuation.period_id == period_id,
                )
            )
            existing = existing_result.scalar_one_or_none()
            if existing:
                existing.item_name = d.get("name")
                existing.category = InventoryCategory.raw_material
                existing.material_id = mat_id
                existing.product_id = None
                existing.crude_product_id = None
                existing.quantity = stock
                existing.unit = d.get("unit") or "kg"
                existing.standard_unit_price = unit_price
                existing.valuation_amount = valuation
                existing.source_system = src_enum
            else:
                db.add(InventoryValuation(
                    period_id=period_id,
                    item_code=d["orig_code"][:30],
                    item_name=d.get("name"),
                    warehouse_name=warehouse_name,
                    category=InventoryCategory.raw_material,
                    material_id=mat_id,
                    quantity=stock,
                    unit=d.get("unit") or "kg",
                    standard_unit_price=unit_price,
                    valuation_amount=valuation,
                    source_system=src_enum,
                    notes=f"1.5原材料在庫 (集約 {d['lot_count']}ロット)",
                ))
            success += 1
        except Exception as e:
            db.add(ImportErrorModel(
                batch_id=batch.id,
                row_number=0,
                error_message=str(e),
                raw_data={"orig_code": d.get("orig_code"), "name": d.get("name")},
            ))
            errors += 1

    # マスタ単価バッチ更新 + 期別単価 (material_standard_costs) も同期
    msc_inserted = 0
    msc_updated = 0
    if pending_master_updates:
        # 当期の既存 MaterialStandardCost を map 化
        msc_res = await db.execute(
            select(MaterialStandardCost).where(
                MaterialStandardCost.period_id == period_id,
                MaterialStandardCost.material_id.in_(list(pending_master_updates.keys())),
            )
        )
        msc_existing: dict[uuid.UUID, MaterialStandardCost] = {
            r.material_id: r for r in msc_res.scalars().all()
        }
        for mat_id, new_price in pending_master_updates.items():
            res = await db.execute(select(Material).where(Material.id == mat_id))
            m = res.scalar_one_or_none()
            if m:
                m.standard_unit_price = new_price
                master_updated += 1
            # MaterialStandardCost も同期 (period_id 別の履歴を持つ)
            existing_msc = msc_existing.get(mat_id)
            if existing_msc is None:
                db.add(MaterialStandardCost(
                    material_id=mat_id,
                    period_id=period_id,
                    unit_cost=new_price,
                    notes="1.5原材料在庫 取込時に補完",
                ))
                msc_inserted += 1
            elif Decimal(str(existing_msc.unit_cost)) != new_price:
                existing_msc.unit_cost = new_price
                msc_updated += 1

    batch.success_rows = success
    batch.error_rows = errors
    batch.status = ImportStatus.completed if errors == 0 else (
        ImportStatus.failed if success == 0 else ImportStatus.completed
    )
    batch.notes = (
        f"materials 新規補完: {new_mat_count}件, SC有り: {sc_priced}件 / SC無し: {no_sc_count}件"
        + (f", マスタ単価更新: {master_updated}件" if update_master_price else "")
        + (f", 期別単価 INS/UPD: {msc_inserted}/{msc_updated}件" if update_master_price else "")
        + (f", 在庫0スキップ: {skipped_zero}件" if skip_zero_stock else "")
    )
    batch.completed_at = datetime.now()

    await db.flush()
    await db.refresh(batch)
    return batch
