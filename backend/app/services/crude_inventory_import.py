"""原液(原体)在庫インポートサービス —
xlsb「2.9原液在庫」シート (xlsx変換後) から原液タンク在庫を取り込み、
inventory_valuations に category=crude_product として登録する。

シート構造:
  A列: 原液コード (1-44-1, 1-XX-Y, 4-XX-YY, 10-X 等。「製造倉庫」等の文字列はセクション見出し)
  B列: 原液名     (44R, 20HI, GIN, 植物用ブレンド 等)
  C列: 番手 / D列: 種類 / E列: 備考
  F列: 受入数量 / G列: 使用数量 / H列: 在庫数量(=評価対象) / I列: タル数

セクション構造:
  Section 1 (Row 2-233): 原液タンク在庫 (1-XX-Y, 4-XX, 10-X 形式) ← 本サービスが取込
  Section 2 (Row 234+):  製造倉庫の半製品 (col1=「製造倉庫」「製造倉庫（サブ）」、col2=短縮コード)
                          → 4.3期末全在庫の半製品レコードと数量重複するためスキップ

crude_products マスタ補完ロジック:
  コード prefix (例 1-44-1 の "1") + name の先頭から CrudeProductType を推定
  vintage_year は code の中央番号 (1-44-1 → 44) を採用 (推定可能な場合のみ)

標準単価:
  CrudeProductStandardCost (period_id一致) があれば採用、無ければ 0
  Excel側に単価列は存在しない
"""

import io
import re
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit import ImportBatch, ImportError as ImportErrorModel, ImportStatus
from app.models.cost import (
    CrudeProductStandardCost,
    InventoryCategory,
    InventoryValuation,
    SourceSystem,
)
from app.models.master import CrudeProduct, CrudeProductType, FiscalPeriod


CRUDE_INVENTORY_SHEET = "2.9原液在庫"

# 原液タンク在庫の倉庫名(全行に固定で付与)
DEFAULT_WAREHOUSE = "原液タンク"

# Section 1 のコード形式: 数字-数字(-数字)*
CODE_PATTERN_SECTION1 = re.compile(r"^\d+-\d+(-\d+)*$")

# vintage 判定: 1-44-1 / 1-44-1-01 → 中央 = 44
VINTAGE_PATTERN = re.compile(r"^(\d+)-(\d+)(-.*)?$")

# name から crude_type 推定 (先頭から最長マッチ優先)
# 例: "44R" → R, "20HI" → HI, "18HI39Rﾘ" → HI (先頭優先)
NAME_TYPE_PATTERNS: list[tuple[re.Pattern, CrudeProductType]] = [
    (re.compile(r"^\d*HIBkai", re.I), CrudeProductType.HIBkai),
    (re.compile(r"^\d*HIpa", re.I), CrudeProductType.HIpa),
    (re.compile(r"^\d*HIB", re.I), CrudeProductType.HIB),
    (re.compile(r"^\d*HIA", re.I), CrudeProductType.HIA),
    (re.compile(r"^\d*HIR", re.I), CrudeProductType.HIR),
    (re.compile(r"^\d*HI", re.I), CrudeProductType.HI),
    (re.compile(r"^\d*PXA", re.I), CrudeProductType.PXA),
    (re.compile(r"^\d*PXM", re.I), CrudeProductType.PX),  # PXM → PX
    (re.compile(r"^\d*PX", re.I), CrudeProductType.PX),
    (re.compile(r"^\d*PE", re.I), CrudeProductType.PE),
    (re.compile(r"^\d*PSA", re.I), CrudeProductType.PXA),  # PSA → PXA 近似
    (re.compile(r"^\d*GA", re.I), CrudeProductType.GA),
    (re.compile(r"^\d*GB", re.I), CrudeProductType.GB),
    (re.compile(r"^\d*GP", re.I), CrudeProductType.GP),
    (re.compile(r"^GIN", re.I), CrudeProductType.R),  # GIN系 → R扱い
    (re.compile(r"^\d*RB", re.I), CrudeProductType.RB),
    (re.compile(r"^\d*RG", re.I), CrudeProductType.RG),
    (re.compile(r"^\d*Rリ"), CrudeProductType.Rri),
    (re.compile(r"^\d*Rﾘ"), CrudeProductType.Rri),  # 半角カナ
    (re.compile(r"^\d*Rﾏ|^\d*Rマ"), CrudeProductType.Rma),
    (re.compile(r"^\d*Rシ|^\d*Rｼ"), CrudeProductType.Rshi),
    (re.compile(r"^\d*R3"), CrudeProductType.R3),
    (re.compile(r"^\d*R2"), CrudeProductType.R2),
    (re.compile(r"^\d*R1"), CrudeProductType.R1),
    (re.compile(r"^\d*RX", re.I), CrudeProductType.RX),
    (re.compile(r"^\d*R(?![A-Za-z])"), CrudeProductType.R),
    (re.compile(r"^\d*MP", re.I), CrudeProductType.MP),
    (re.compile(r"^\d*LPA", re.I), CrudeProductType.LPA),
    (re.compile(r"^\d*LP", re.I), CrudeProductType.LP),
    (re.compile(r"^\d*FEB", re.I), CrudeProductType.FEB),
    (re.compile(r"^\d*FB", re.I), CrudeProductType.FB),
    (re.compile(r"^\d*BM", re.I), CrudeProductType.BM),
    (re.compile(r"^\d*B(?![A-Za-z])"), CrudeProductType.B),
    (re.compile(r"^植物用|^ﾐｬﾝﾏｰ"), CrudeProductType.plant),
]


def _to_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _is_section1_data_code(code_str: str) -> bool:
    """Section1 のデータ行(原液コード形式)かどうか。
    倉庫名や見出しは除外。
    """
    if not code_str:
        return False
    if "倉庫" in code_str or "小計" in code_str or "合計" in code_str:
        return False
    return bool(CODE_PATTERN_SECTION1.match(code_str))


def _infer_crude_type(name: str | None, code: str) -> CrudeProductType:
    """name → CrudeProductType 推定。マッチしなければ other。"""
    if not name:
        return CrudeProductType.other
    n = str(name).strip()
    for pat, t in NAME_TYPE_PATTERNS:
        if pat.match(n):
            return t
    return CrudeProductType.other


def _infer_vintage_year(code: str) -> int | None:
    """code から vintage 年度を推定。1-44-1 → 44。
    対象は prefix '1' (R系・HI系) のみ。それ以外は None。
    """
    m = VINTAGE_PATTERN.match(code)
    if not m:
        return None
    prefix, mid, _ = m.groups()
    if prefix != "1":
        return None
    try:
        v = int(mid)
        if 10 <= v <= 99:  # 妥当な期年範囲
            return v
    except ValueError:
        pass
    return None


def _parse_crude_inventory_xlsx(content: bytes, sheet_name: str = CRUDE_INVENTORY_SHEET) -> list[dict]:
    """2.9原液在庫シートをパース。Section1のデータ行のみ返す。"""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    rows: list[dict] = []
    skipped_section2 = 0
    for r_idx, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            continue  # ヘッダ
        if not row_values or row_values[0] is None:
            continue
        code = str(row_values[0]).strip()
        if not code:
            continue
        # Section 2 (倉庫名がcol1) はスキップ
        if "倉庫" in code:
            skipped_section2 += 1
            continue
        if not _is_section1_data_code(code):
            continue

        name = str(row_values[1]).strip() if len(row_values) > 1 and row_values[1] is not None else None
        rows.append({
            "row_number": r_idx,
            "code": code,
            "name": name,
            "banshu": str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] is not None else None,
            "kind":   str(row_values[3]).strip() if len(row_values) > 3 and row_values[3] is not None else None,
            "memo":   str(row_values[4]).strip() if len(row_values) > 4 and row_values[4] is not None else None,
            "stock_raw": row_values[7] if len(row_values) > 7 else None,
            "taru":      row_values[8] if len(row_values) > 8 else None,
        })
    wb.close()
    return rows


async def process_crude_inventory_import(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    period_id: uuid.UUID,
    sheet_name: str = CRUDE_INVENTORY_SHEET,
    source_system: str = "manual",
    warehouse_name: str = DEFAULT_WAREHOUSE,
    delete_existing: bool = True,
    skip_zero_stock: bool = False,
) -> ImportBatch:
    """2.9原液在庫シートを取り込み、inventory_valuations にUPSERTする。

    crude_products マスタに無いコードは新規INSERT (vintage_year, crude_type を推定)。
    delete_existing=True の場合、同一period × category=crude_product × warehouse の
    既存レコードを削除してから登録。
    """
    batch = ImportBatch(
        file_name=filename,
        source_system=source_system,
        status=ImportStatus.processing,
        period_id=period_id,
        total_rows=0,
        success_rows=0,
        error_rows=0,
        started_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()

    period_result = await db.execute(
        select(FiscalPeriod).where(FiscalPeriod.id == period_id)
    )
    period = period_result.scalar_one_or_none()
    if not period:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"指定された会計期間が見つかりません: {period_id}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    # Parse
    try:
        rows = _parse_crude_inventory_xlsx(file_content, sheet_name)
    except Exception as e:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"ファイルパースエラー: {e}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    batch.total_rows = len(rows)

    # 既存削除 (period × category=crude_product × warehouse)
    if delete_existing:
        from sqlalchemy import delete
        await db.execute(
            delete(InventoryValuation).where(
                InventoryValuation.period_id == period_id,
                InventoryValuation.category == InventoryCategory.crude_product,
                InventoryValuation.warehouse_name == warehouse_name,
            )
        )

    # crude_products 既存マップ (code → id)
    crude_result = await db.execute(select(CrudeProduct.code, CrudeProduct.id))
    crude_map: dict[str, uuid.UUID] = {r[0]: r[1] for r in crude_result.all()}

    # CrudeProductStandardCost マップ (crude_product_id → unit_cost)
    sc_result = await db.execute(
        select(CrudeProductStandardCost.crude_product_id, CrudeProductStandardCost.unit_cost)
        .where(CrudeProductStandardCost.period_id == period_id)
    )
    crude_sc_map: dict[uuid.UUID, Decimal] = {r[0]: r[1] for r in sc_result.all()}

    src_enum = SourceSystem(source_system) if source_system in SourceSystem._value2member_map_ else SourceSystem.manual

    success = 0
    errors = 0
    new_crude_count = 0
    skipped_zero = 0

    for row in rows:
        try:
            code = row["code"]
            stock = _to_decimal(row["stock_raw"])
            # 微小な浮動小数点誤差(2.84e-14等)をゼロ扱い
            if stock is not None and abs(stock) < Decimal("0.001"):
                stock = Decimal("0")
            if stock is None:
                stock = Decimal("0")
            if skip_zero_stock and stock == 0:
                skipped_zero += 1
                continue

            # crude_products lookup (or create)
            crude_id = crude_map.get(code)
            if crude_id is None:
                crude_type = _infer_crude_type(row.get("name"), code)
                vintage = _infer_vintage_year(code)
                new_crude = CrudeProduct(
                    code=code[:20],
                    name=(row.get("name") or code)[:200],
                    crude_type=crude_type,
                    vintage_year=vintage,
                    is_blend=False,
                    unit="kg",
                    is_active=True,
                    notes="2.9原液在庫 取込時に補完",
                )
                db.add(new_crude)
                await db.flush()
                crude_id = new_crude.id
                crude_map[code] = crude_id
                new_crude_count += 1

            unit_price = crude_sc_map.get(crude_id, Decimal("0"))
            valuation = stock * unit_price

            # UPSERT (item_code × warehouse × period)
            existing_result = await db.execute(
                select(InventoryValuation).where(
                    InventoryValuation.item_code == code,
                    InventoryValuation.warehouse_name == warehouse_name,
                    InventoryValuation.period_id == period_id,
                )
            )
            existing = existing_result.scalar_one_or_none()
            if existing:
                existing.item_name = row.get("name")
                existing.category = InventoryCategory.crude_product
                existing.crude_product_id = crude_id
                existing.product_id = None
                existing.material_id = None
                existing.quantity = stock
                existing.unit = "kg"
                existing.standard_unit_price = unit_price
                existing.valuation_amount = valuation
                existing.source_system = src_enum
            else:
                db.add(InventoryValuation(
                    period_id=period_id,
                    item_code=code,
                    item_name=row.get("name"),
                    warehouse_name=warehouse_name,
                    category=InventoryCategory.crude_product,
                    crude_product_id=crude_id,
                    quantity=stock,
                    unit="kg",
                    standard_unit_price=unit_price,
                    valuation_amount=valuation,
                    source_system=src_enum,
                    notes=f'2.9原液在庫 (banshu={row.get("banshu") or ""}, kind={row.get("kind") or ""})',
                ))
            success += 1
        except Exception as e:
            db.add(ImportErrorModel(
                batch_id=batch.id,
                row_number=row.get("row_number", 0),
                error_message=str(e),
                raw_data={k: (str(v) if v is not None else None) for k, v in row.items() if k != "row_number"},
            ))
            errors += 1

    batch.success_rows = success
    batch.error_rows = errors
    batch.status = ImportStatus.completed if errors == 0 else (
        ImportStatus.failed if success == 0 else ImportStatus.completed
    )
    batch.notes = (
        f"crude_products 新規補完: {new_crude_count}件"
        + (f", 在庫0スキップ: {skipped_zero}件" if skip_zero_stock else "")
    )
    batch.completed_at = datetime.now()

    await db.flush()
    await db.refresh(batch)
    return batch
