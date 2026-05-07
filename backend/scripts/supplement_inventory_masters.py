"""未登録 item_code を products マスタに補完し、StandardCost を整備する。

在庫評価 (4.3期末全在庫) で発生した「マスタ未登録 → product_id=NULL」
レコードを解消するためのワンショット補完スクリプト。

処理:
  Step1: orphan (product_id=NULL, category=製品/半製品/商品) のユニーク
         item_code を products テーブルへINSERT
            category=product       → product_type=in_house_product_dept
            category=semi_finished → product_type=semi_finished
            category=merchandise   → product_type=outsourced
  Step2: 4.3期末全在庫.xlsx を再インポート (UPSERT で product_id 再リンク)
  Step3: Excel L列 (在庫単価) に値がある分は StandardCost (period_id) へINSERT
  Step4: 標準原価_製品_2603v5-2.xlsx「製品標準原価」シートの 39期SC col28
         から StandardCost を補完 (主に製品110件のうち85件)
  Step5: recalculate_valuation_amounts で全体の評価金額を更新

使い方:
  cd backend
  python -m scripts.supplement_inventory_masters \\
      --period-id 1c533e58-25a8-4e87-9693-e07eb202611a \\
      --inventory-xlsx /tmp/inv-convert/4_3_期末全在庫_38ki_M01.xlsx \\
      --sc-xlsx 'docs/reference/標準原価_製品_2603v5ー2.xlsx'

注意:
  - .xlsb 形式の場合は事前に 4.3期末全在庫 シートのみ openpyxl 読込み可能な
    .xlsx に変換すること(`scripts/convert_xlsb_sheet.py` 参照)
  - 補完した products には notes='inventory orphan 補完 ...' を付与し追跡可能
  - 補完した standard_costs には notes='4.3期末全在庫 L列単価 (補完)' または
    '製品標準原価シート (39期SC col28、補完)' を付与
"""
import argparse
import asyncio
import os
import sys
import uuid
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

# backend ルートを sys.path に追加
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from sqlalchemy import select, func
from app.db.session import async_session_factory
from app.models.cost import (
    InventoryCategory, InventoryValuation, StandardCost,
)
from app.models.master import Product, ProductType
from app.services.inventory_import import process_inventory_import
from app.services.inventory_valuation import recalculate_valuation_amounts


CATEGORY_TO_PRODUCT_TYPE = {
    InventoryCategory.product:       ProductType.in_house_product_dept,
    InventoryCategory.semi_finished: ProductType.semi_finished,
    InventoryCategory.merchandise:   ProductType.outsourced,
}


def parse_inventory_index(xlsx_path: str) -> dict[str, dict]:
    """期末全在庫 Excel から item_code → {name, unit, l_price} を構築。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    idx: dict[str, dict] = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri == 1:
            continue
        a = row[0]
        if a is None:
            continue
        code = str(a).strip()
        name = row[3] if len(row) > 3 else None
        unit = row[5] if len(row) > 5 else None
        L = row[11] if len(row) > 11 else None
        if code not in idx:
            idx[code] = {'name': name, 'unit': unit or '個', 'l_price': None}
        d = idx[code]
        if L is not None and d['l_price'] is None:
            try:
                v = Decimal(str(L))
                if v > 0:
                    d['l_price'] = v
            except Exception:
                pass
    wb.close()
    return idx


def parse_sc_sheet_index(xlsx_path: str, sheet_name: str = '製品標準原価') -> dict[str, Decimal]:
    """標準原価_製品_2603 系の「製品標準原価」シートから
    code → 39期SC単価 (col 28) を抽出。
    シート構造: col3=商品コード, col28=39期標準原価円/個(計)。
    ヘッダ行は1-5、データは6行目以降。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    m: dict[str, Decimal] = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ri <= 5:
            continue
        code_v = row[2] if len(row) > 2 else None
        if code_v is None:
            continue
        try:
            code = str(int(float(code_v)))
        except (ValueError, TypeError):
            continue
        sc_39 = row[27] if len(row) > 27 else None
        if sc_39 is None:
            continue
        try:
            v = Decimal(str(sc_39))
            if v > 0:
                m[code] = v
        except Exception:
            pass
    wb.close()
    return m


async def step1_insert_missing_products(db, period_id: uuid.UUID, excel_idx: dict[str, dict]) -> int:
    """orphan の item_code を products にINSERT。"""
    res = await db.execute(select(Product.code))
    existing = {r[0] for r in res.all()}

    res = await db.execute(
        select(InventoryValuation.item_code, InventoryValuation.category,
               InventoryValuation.item_name, InventoryValuation.unit)
        .where(
            InventoryValuation.period_id == period_id,
            InventoryValuation.category.in_([
                InventoryCategory.product, InventoryCategory.semi_finished,
                InventoryCategory.merchandise,
            ]),
            InventoryValuation.product_id.is_(None),
        ).distinct()
    )
    rows = res.all()

    inserted = 0
    skipped = 0
    by_cat: dict[tuple[str, str], int] = defaultdict(int)
    for item_code, cat, item_name, unit in rows:
        if item_code in existing:
            skipped += 1
            continue
        ex = excel_idx.get(item_code, {})
        name = ex.get('name') or item_name or item_code
        u = ex.get('unit') or unit or '個'
        ptype = CATEGORY_TO_PRODUCT_TYPE.get(cat, ProductType.outsourced)
        db.add(Product(
            code=item_code,
            name=str(name)[:200],
            product_type=ptype,
            unit=str(u)[:10],
            standard_lot_size=Decimal('1'),
            is_active=True,
            notes='inventory orphan 補完 (4.3期末全在庫)',
        ))
        existing.add(item_code)
        inserted += 1
        by_cat[(cat.value, ptype.value)] += 1

    await db.flush()
    print(f"\n=== Step1: products INSERT ===")
    print(f"  inserted: {inserted}, skipped(既存): {skipped}")
    for (c, t), n in sorted(by_cat.items()):
        print(f"    {c} → {t}: {n}件")
    return inserted


async def step2_reimport(db, period_id: uuid.UUID, xlsx_path: str) -> dict:
    """4.3期末全在庫 を再インポートして product_id を再リンク。"""
    with open(xlsx_path, 'rb') as f:
        content = f.read()
    batch = await process_inventory_import(
        db=db,
        file_content=content,
        filename=Path(xlsx_path).name,
        period_id=period_id,
        source_system='manual',
    )
    print(f"\n=== Step2: re-import ===")
    print(f"  status={batch.status}, total={batch.total_rows}, "
          f"success={batch.success_rows}, errors={batch.error_rows}")
    return {'total': batch.total_rows, 'success': batch.success_rows, 'errors': batch.error_rows}


async def step3_sc_from_l_column(db, period_id: uuid.UUID, excel_idx: dict[str, dict]) -> int:
    """期末全在庫 Excel の L列単価 を StandardCost に登録。"""
    res = await db.execute(
        select(Product.id, Product.code).where(
            Product.notes.like('inventory orphan 補完%')
        )
    )
    new_products = {r[1]: r[0] for r in res.all()}

    res = await db.execute(
        select(StandardCost.product_id).where(StandardCost.period_id == period_id)
    )
    existing_pids = {r[0] for r in res.all()}

    inserted = 0
    skipped_no_price = 0
    skipped_existing = 0
    for code, pid in new_products.items():
        if pid in existing_pids:
            skipped_existing += 1
            continue
        ex = excel_idx.get(code, {})
        l_price = ex.get('l_price')
        if l_price is None:
            skipped_no_price += 1
            continue
        db.add(StandardCost(
            product_id=pid,
            period_id=period_id,
            crude_product_cost=Decimal('0'),
            packaging_cost=Decimal('0'),
            labor_cost=Decimal('0'),
            overhead_cost=Decimal('0'),
            outsourcing_cost=Decimal('0'),
            total_cost=l_price,
            unit_cost=l_price,
            lot_size=Decimal('1'),
            notes='4.3期末全在庫 L列単価 (補完)',
        ))
        inserted += 1
    await db.flush()
    print(f"\n=== Step3: StandardCost INSERT (Excel L列) ===")
    print(f"  inserted: {inserted}, skipped(no L price): {skipped_no_price}, "
          f"skipped(existing): {skipped_existing}")
    return inserted


async def step4_sc_from_external_sheet(db, period_id: uuid.UUID, sc_xlsx: str) -> int:
    """標準原価_製品_2603v5-2.xlsx の「製品標準原価」シートから補完。"""
    sc_map = parse_sc_sheet_index(sc_xlsx)
    print(f"\n  外部SCシート: {len(sc_map)} entries from {Path(sc_xlsx).name}")

    res = await db.execute(
        select(Product.id, Product.code).where(
            Product.notes.like('inventory orphan 補完%')
        )
    )
    new_products = {r[1]: r[0] for r in res.all()}

    res = await db.execute(
        select(StandardCost.product_id).where(StandardCost.period_id == period_id)
    )
    existing_pids = {r[0] for r in res.all()}

    inserted = 0
    skipped_no_match = 0
    skipped_existing = 0
    for code, pid in new_products.items():
        if pid in existing_pids:
            skipped_existing += 1
            continue
        v = sc_map.get(code)
        if v is None:
            skipped_no_match += 1
            continue
        db.add(StandardCost(
            product_id=pid,
            period_id=period_id,
            crude_product_cost=Decimal('0'),
            packaging_cost=Decimal('0'),
            labor_cost=Decimal('0'),
            overhead_cost=Decimal('0'),
            outsourcing_cost=Decimal('0'),
            total_cost=v,
            unit_cost=v,
            lot_size=Decimal('1'),
            notes='製品標準原価シート (39期SC col28、補完)',
        ))
        inserted += 1
    await db.flush()
    print(f"=== Step4: StandardCost INSERT (外部SCシート) ===")
    print(f"  inserted: {inserted}, skipped(no match): {skipped_no_match}, "
          f"skipped(existing): {skipped_existing}")
    return inserted


async def step5_recalculate(db, period_id: uuid.UUID) -> int:
    n = await recalculate_valuation_amounts(db, period_id)
    print(f"\n=== Step5: recalculate ===")
    print(f"  updated: {n}")
    return n


async def snapshot(db, period_id: uuid.UUID, label: str):
    total = (await db.execute(select(func.count()).where(
        InventoryValuation.period_id == period_id))).scalar()
    by_cat = await db.execute(
        select(
            InventoryValuation.category,
            func.count().label('records'),
            func.count().filter(InventoryValuation.standard_unit_price > 0).label('priced'),
            func.count().filter(InventoryValuation.product_id.is_(None)).label('orphan'),
            func.sum(InventoryValuation.valuation_amount).label('sum_val'),
        )
        .where(InventoryValuation.period_id == period_id)
        .group_by(InventoryValuation.category)
    )
    sum_total = (await db.execute(select(func.sum(InventoryValuation.valuation_amount)).where(
        InventoryValuation.period_id == period_id))).scalar()
    print(f"\n--- {label} (total={total}, sum_valuation={sum_total:,.0f}) ---")
    print(f"  {'category':<15} {'records':>8} {'priced':>8} {'orphan':>8}  {'sum_val':>18}")
    for cat, rec, priced, orphan, sv in by_cat.all():
        print(f"  {cat.value:<15} {rec:>8} {priced:>8} {orphan:>8}  {float(sv or 0):>18,.0f}")


async def run(period_id: uuid.UUID, inventory_xlsx: str, sc_xlsx: str | None):
    excel_idx = parse_inventory_index(inventory_xlsx)
    print(f"在庫Excel index: {len(excel_idx)} unique codes from {Path(inventory_xlsx).name}")

    async with async_session_factory() as db:
        await snapshot(db, period_id, 'BEFORE')
        await step1_insert_missing_products(db, period_id, excel_idx)
        await step2_reimport(db, period_id, inventory_xlsx)
        await step3_sc_from_l_column(db, period_id, excel_idx)
        if sc_xlsx:
            await step4_sc_from_external_sheet(db, period_id, sc_xlsx)
        await step5_recalculate(db, period_id)
        await db.commit()
        await snapshot(db, period_id, 'AFTER')


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--period-id', required=True, help='対象会計期間のUUID')
    parser.add_argument('--inventory-xlsx', required=True, help='4.3期末全在庫 .xlsx ファイルパス')
    parser.add_argument('--sc-xlsx', default=None,
                        help='標準原価_製品_2603v5-2.xlsx ファイルパス (Step4で使用、任意)')
    args = parser.parse_args()

    os.environ.setdefault('SECRET_KEY', 'dev')
    os.environ.setdefault('ANTHROPIC_API_KEY', 'dummy')

    asyncio.run(run(uuid.UUID(args.period_id), args.inventory_xlsx, args.sc_xlsx))


if __name__ == '__main__':
    main()
