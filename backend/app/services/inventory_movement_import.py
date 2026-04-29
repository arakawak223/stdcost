"""Product inventory movement import service —
Excel「製品増減内訳表」シートから製品の38期数量増減を InventoryMovement に取り込む。

シート構造（製品増減内訳表）:
  A列: 区分(半/品/造/内/外内/外/他), C列: 商品コード, D列: 商品名
  数量情報 (col 42-54):
    42: 期首(37期末), 43: 生産, 44: 販売, 45: 外注支給,
    46: 販促DM, 47: 販促, 48: 試験研究費, 49: 接待交際費,
    50: 寄付金, 51: 広告宣伝費, 52: 在庫調整, 53: 振替, 54: 期末

MovementType マッピング(テスト用、既存値再利用):
  生産           → finished_goods (受入)
  販促DM/販促     → promotion (払出)
  試験研究費      → research (払出)
  その他出庫     → adjustment (販売/外注支給/接待/寄付/広告/在庫調整/振替)
"""

import io
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit import ImportBatch, ImportError as ImportErrorModel, ImportStatus
from app.models.cost import (
    InventoryMovement,
    MovementType,
    SourceSystem,
    StandardCost,
)
from app.models.master import CostCenter, FiscalPeriod, Product

PRODUCT_MOVEMENT_SHEET = "製品増減内訳表"

# (Excel列番号, MovementType, ラベル) のリスト
COLUMN_MOVEMENT_MAP: list[tuple[int, MovementType, str]] = [
    (43, MovementType.finished_goods, "生産"),
    (44, MovementType.adjustment, "販売"),
    (45, MovementType.adjustment, "外注支給"),
    (46, MovementType.promotion, "販促DM"),
    (47, MovementType.promotion, "販促"),
    (48, MovementType.research, "試験研究費"),
    (49, MovementType.adjustment, "接待交際費"),
    (50, MovementType.adjustment, "寄付金"),
    (51, MovementType.adjustment, "広告宣伝費"),
    (52, MovementType.adjustment, "在庫調整"),
    (53, MovementType.adjustment, "振替"),
]

VALID_KIND_VALUES = {"半", "品", "造", "内", "外内", "外", "他"}


def _to_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        d = Decimal(str(v))
        return d if d != 0 else None
    except (InvalidOperation, ValueError):
        return None


def _parse_movement_xlsx(content: bytes, sheet_name: str = PRODUCT_MOVEMENT_SHEET) -> list[dict]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    rows: list[dict] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < 6:
            continue  # ヘッダ行
        if not row:
            continue
        kind = row[0] if len(row) > 0 else None
        if kind not in VALID_KIND_VALUES:
            continue
        code = row[2] if len(row) > 2 else None
        if code is None:
            continue
        item_code = str(code).strip()

        # col 42=期首, col 54=期末, col 43-53=movements (Pythonは0-indexed)
        movements = []
        for col_num, mv_type, label in COLUMN_MOVEMENT_MAP:
            idx = col_num - 1
            v = row[idx] if len(row) > idx else None
            qty = _to_decimal(v)
            if qty is None:
                continue
            movements.append({"movement_type": mv_type, "label": label, "quantity": qty})

        if not movements:
            continue

        rows.append({
            "row_number": r_idx,
            "item_code": item_code,
            "item_name": str(row[3]) if len(row) > 3 and row[3] else None,
            "movements": movements,
        })
    wb.close()
    return rows


async def _get_product_cost_center(db: AsyncSession) -> uuid.UUID | None:
    """製品課の cost_center_id を取得（無ければ最初のcost_center）。"""
    from app.models.master import CostCenterType
    result = await db.execute(
        select(CostCenter.id).where(CostCenter.center_type == CostCenterType.product).limit(1)
    )
    cc_id = result.scalar_one_or_none()
    if cc_id:
        return cc_id
    # フォールバック
    result = await db.execute(select(CostCenter.id).limit(1))
    return result.scalar_one_or_none()


async def process_product_movement_import(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
    period_id: uuid.UUID,
    sheet_name: str = PRODUCT_MOVEMENT_SHEET,
    source_system: str = "manual",
    delete_existing: bool = True,
) -> ImportBatch:
    """製品増減内訳表シートを取り込み、各列を MovementType にマップして InventoryMovement に登録。

    delete_existing=True の場合、同一期間の既存 InventoryMovement (該当ソース) を削除してから登録。
    """
    batch = ImportBatch(
        file_name=filename,
        source_system=source_system,
        status=ImportStatus.processing,
        period_id=period_id,
        total_rows=0,
        success_rows=0,
        error_rows=0,
        started_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()

    # 期間取得
    period_result = await db.execute(
        select(FiscalPeriod).where(FiscalPeriod.id == period_id)
    )
    period = period_result.scalar_one_or_none()
    if not period:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"指定された会計期間が見つかりません: {period_id}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    movement_date = period.end_date

    # 製品マスタ map (code → product_id)
    prod_result = await db.execute(select(Product.code, Product.id))
    prod_map: dict[str, uuid.UUID] = {row[0]: row[1] for row in prod_result.all()}

    # 標準原価 map (product_id → unit_cost)
    std_result = await db.execute(
        select(StandardCost.product_id, StandardCost.unit_cost).where(
            StandardCost.period_id == period_id
        )
    )
    std_map: dict[uuid.UUID, Decimal] = {row[0]: row[1] for row in std_result.all()}

    # cost_center_id
    cc_id = await _get_product_cost_center(db)
    if not cc_id:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = "製品課の cost_center が見つかりません"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    # Parse
    try:
        rows = _parse_movement_xlsx(file_content, sheet_name)
    except Exception as e:
        batch.status = ImportStatus.failed
        batch.completed_at = datetime.now()
        batch.notes = f"ファイルパースエラー: {e}"
        db.add(ImportErrorModel(batch_id=batch.id, row_number=0, error_message=batch.notes))
        await db.flush()
        await db.refresh(batch)
        return batch

    batch.total_rows = len(rows)

    # 既存削除（同一期間×same source）
    if delete_existing:
        from sqlalchemy import delete
        await db.execute(
            delete(InventoryMovement).where(
                InventoryMovement.period_id == period_id,
                InventoryMovement.product_id.is_not(None),
                InventoryMovement.source_system == SourceSystem(source_system),
            )
        )

    src_enum = SourceSystem(source_system) if source_system in SourceSystem._value2member_map_ else SourceSystem.manual

    success = 0
    errors = 0
    skipped_unmapped = 0

    for row in rows:
        try:
            product_id = prod_map.get(row["item_code"])
            if not product_id:
                skipped_unmapped += 1
                continue
            unit_cost = std_map.get(product_id, Decimal("0"))

            for mv in row["movements"]:
                qty = mv["quantity"]
                # 符号統一: 全て正の絶対値で登録（受入/払出はMovementTypeで判別）
                qty_abs = abs(qty)
                total = qty_abs * unit_cost
                db.add(InventoryMovement(
                    product_id=product_id,
                    cost_center_id=cc_id,
                    period_id=period_id,
                    movement_type=mv["movement_type"],
                    movement_date=movement_date,
                    quantity=qty_abs,
                    unit_cost=unit_cost,
                    total_cost=total,
                    source_system=src_enum,
                    notes=f'{mv["label"]} (38期、製品増減内訳表)',
                ))
            success += 1
        except Exception as e:
            db.add(ImportErrorModel(
                batch_id=batch.id,
                row_number=row.get("row_number", 0),
                error_message=str(e),
                raw_data={"item_code": row["item_code"]},
            ))
            errors += 1

    batch.success_rows = success
    batch.error_rows = errors
    batch.status = ImportStatus.completed if errors == 0 else (
        ImportStatus.failed if success == 0 else ImportStatus.completed
    )
    batch.notes = f"マスタ未登録でスキップ: {skipped_unmapped}件"
    batch.completed_at = datetime.now()
    await db.flush()
    await db.refresh(batch)
    return batch
